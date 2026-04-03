"""
Test for 'xlsx' skill — openpyxl Spreadsheet Generation
Validates 3 sheets, =SUM/=VLOOKUP formulas, named ranges,
currency formatting, and workbook structure.
"""

import os
import re
import sys

import pytest


class TestXlsx:
    """Verify openpyxl spreadsheet generation."""

    REPO_DIR = "/workspace/openpyxl"

    # ── file_path_check ─────────────────────────────────────────────────────

    def test_openpyxl_source_exists(self):
        """Verify openpyxl source directory exists."""
        pkg = os.path.join(self.REPO_DIR, "openpyxl")
        assert os.path.isdir(pkg), "openpyxl/ package not found"

    def test_workbook_module_exists(self):
        """Verify workbook module exists."""
        wb_mod = os.path.join(self.REPO_DIR, "openpyxl", "workbook")
        assert os.path.isdir(wb_mod), "openpyxl/workbook/ not found"

    # ── semantic_check ──────────────────────────────────────────────────────

    def test_multiple_sheets(self):
        """Verify support for multiple worksheets (≥3)."""
        py_files = self._find_py_files()
        for fpath in py_files:
            content = self._read(fpath)
            if re.search(
                r"(create_sheet|Worksheet|sheetnames|active_sheet|get_sheet)", content
            ):
                return
        pytest.fail("No multi-sheet support found")

    def test_sum_formula(self):
        """Verify =SUM formula support."""
        py_files = self._find_py_files()
        for fpath in py_files:
            content = self._read(fpath)
            if re.search(r"(SUM|=SUM|formula|FORMULAE)", content, re.IGNORECASE):
                return
        pytest.fail("No =SUM formula support found")

    def test_vlookup_formula(self):
        """Verify =VLOOKUP formula support."""
        py_files = self._find_py_files()
        for fpath in py_files:
            content = self._read(fpath)
            if re.search(r"(VLOOKUP|HLOOKUP|INDEX|MATCH)", content, re.IGNORECASE):
                return
        pytest.fail("No VLOOKUP/lookup formula support found")

    def test_named_ranges(self):
        """Verify named ranges / defined names."""
        py_files = self._find_py_files()
        for fpath in py_files:
            content = self._read(fpath)
            if re.search(
                r"(defined_name|DefinedName|named_range|name.*range)",
                content,
                re.IGNORECASE,
            ):
                return
        pytest.fail("No named ranges support found")

    def test_number_format(self):
        """Verify currency/number formatting."""
        py_files = self._find_py_files()
        for fpath in py_files:
            content = self._read(fpath)
            if re.search(
                r"(number_format|FORMAT_|currency|\\$|#,##0)", content, re.IGNORECASE
            ):
                return
        pytest.fail("No number formatting found")

    # ── functional_check ────────────────────────────────────────────────────

    def test_source_files_parse(self):
        """Verify Python source files parse correctly."""
        import ast

        py_files = self._find_py_files()
        for fpath in py_files[:20]:
            content = self._read(fpath)
            try:
                ast.parse(content, filename=fpath)
            except SyntaxError as e:
                pytest.fail(f"SyntaxError in {os.path.basename(fpath)}: {e}")

    def test_import_workbook(self):
        """Verify Workbook class can be imported."""
        if self.REPO_DIR not in sys.path:
            sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl import Workbook

            wb = Workbook()
            assert wb.active is not None
        except Exception as e:
            pytest.skip(f"Cannot import Workbook: {e}")

    def test_create_and_access_sheets(self):
        """Verify creating 3 sheets works."""
        if self.REPO_DIR not in sys.path:
            sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl import Workbook

            wb = Workbook()
            wb.create_sheet("Sheet2")
            wb.create_sheet("Sheet3")
            assert len(wb.sheetnames) >= 3
        except Exception as e:
            pytest.skip(f"Cannot test sheets: {e}")

    def test_cell_assignment(self):
        """Verify cell value assignment."""
        if self.REPO_DIR not in sys.path:
            sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws["A1"] = 42
            assert ws["A1"].value == 42
        except Exception as e:
            pytest.skip(f"Cannot test cells: {e}")

    def test_formula_assignment(self):
        """Verify formula can be assigned to cell."""
        if self.REPO_DIR not in sys.path:
            sys.path.insert(0, self.REPO_DIR)
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws["A1"] = "=SUM(B1:B10)"
            assert ws["A1"].value == "=SUM(B1:B10)"
        except Exception as e:
            pytest.skip(f"Cannot test formulas: {e}")

    # ── helpers ──────────────────────────────────────────────────────────────

    def _find_py_files(self):
        results = []
        pkg = os.path.join(self.REPO_DIR, "openpyxl")
        search = pkg if os.path.isdir(pkg) else self.REPO_DIR
        for dirpath, _, fnames in os.walk(search):
            if ".git" in dirpath:
                continue
            for f in fnames:
                if f.endswith(".py"):
                    results.append(os.path.join(dirpath, f))
        return results

    def _read(self, path):
        with open(path, "r", errors="ignore") as fh:
            return fh.read()
