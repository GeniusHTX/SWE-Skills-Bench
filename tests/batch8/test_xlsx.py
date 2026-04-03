"""
Test for 'xlsx' skill — Excel Report Generator
Validates that the Agent implemented an openpyxl-based Excel report generator
with chart builder, worksheet formatter, bold headers, and currency formatting.
"""

import os
import re
import sys

import pytest


class TestXlsx:
    """Verify Excel report generator implementation."""

    REPO_DIR = "/workspace/openpyxl"

    @staticmethod
    def _read(path: str) -> str:
        try:
            with open(path, "r", errors="ignore") as fh:
                return fh.read()
        except OSError:
            return ""

    def _find_file(self, *candidates):
        for c in candidates:
            p = os.path.join(self.REPO_DIR, c)
            if os.path.isfile(p):
                return p
        return None

    # ── file_path_check ─────────────────────────────────────────────

    def test_generator_module_exists(self):
        """Verify xlsx/generator.py or excel_report_generator.py exists."""
        candidates = ("xlsx/generator.py", "excel_report_generator.py")
        found = any(
            os.path.isfile(os.path.join(self.REPO_DIR, c)) for c in candidates)
        assert found, f"Missing: none of {candidates} found"

    def test_chart_builder_module_exists(self):
        """Verify xlsx/chart_builder.py or chart_builder.py exists."""
        candidates = ("xlsx/chart_builder.py", "chart_builder.py")
        found = any(
            os.path.isfile(os.path.join(self.REPO_DIR, c)) for c in candidates)
        assert found, f"Missing: none of {candidates} found"

    def test_formatter_module_exists(self):
        """Verify xlsx/formatter.py or worksheet_formatter.py exists."""
        candidates = ("xlsx/formatter.py", "worksheet_formatter.py")
        found = any(
            os.path.isfile(os.path.join(self.REPO_DIR, c)) for c in candidates)
        assert found, f"Missing: none of {candidates} found"

    # ── semantic_check ──────────────────────────────────────────────

    def test_openpyxl_workbook_instantiation(self):
        """Verify openpyxl.Workbook() is instantiated in generator module."""
        path = self._find_file("xlsx/generator.py", "excel_report_generator.py")
        assert path, "Generator module not found"
        content = self._read(path)
        found = any(kw in content for kw in ("openpyxl.Workbook", "Workbook()"))
        assert found, "Workbook instantiation not found"

    def test_chart_class_imported(self):
        """Verify BarChart or LineChart is imported from openpyxl.chart."""
        path = self._find_file("xlsx/chart_builder.py", "chart_builder.py")
        assert path, "Chart builder module not found"
        content = self._read(path)
        found = any(kw in content for kw in ("BarChart", "LineChart"))
        assert found, "BarChart or LineChart not found"

    def test_bold_font_in_formatter(self):
        """Verify Font(bold=True) is applied to header cells."""
        path = self._find_file("xlsx/formatter.py", "worksheet_formatter.py")
        assert path, "Formatter module not found"
        content = self._read(path)
        assert "bold=True" in content, "bold=True not found"
        found = any(kw in content for kw in ("Font(bold", "Font( bold"))
        assert found, "Font(bold=True) not found in formatter"

    def test_currency_number_format(self):
        """Verify '#,##0.00' number format is assigned to currency cells."""
        path = self._find_file("xlsx/formatter.py", "worksheet_formatter.py")
        assert path, "Formatter module not found"
        content = self._read(path)
        assert "#,##0.00" in content, "#,##0.00 number format not found"

    # ── functional_check (import) ───────────────────────────────────

    def _skip_unless_importable(self):
        if not os.path.isdir(self.REPO_DIR):
            pytest.skip("Repo dir does not exist")
        if self.REPO_DIR not in sys.path:
            sys.path.insert(0, self.REPO_DIR)

    def test_generate_creates_valid_xlsx(self):
        """generate(sample_data, tmp_path) creates a valid xlsx file."""
        self._skip_unless_importable()
        import tempfile
        try:
            import openpyxl
            from xlsx.generator import ExcelReportGenerator
        except Exception as exc:
            pytest.skip(f"Cannot import: {exc}")
        data = [{"item": "A", "amount": 1000.0},
                {"item": "B", "amount": 2500.0}]
        out = tempfile.mktemp(suffix=".xlsx")
        try:
            ExcelReportGenerator().generate(data, out)
            wb = openpyxl.load_workbook(out)
            assert wb is not None
        finally:
            if os.path.exists(out):
                os.unlink(out)

    def test_workbook_has_summary_sheet(self):
        """Generated workbook contains 'Summary' in its sheetnames."""
        self._skip_unless_importable()
        import tempfile
        try:
            import openpyxl
            from xlsx.generator import ExcelReportGenerator
        except Exception as exc:
            pytest.skip(f"Cannot import: {exc}")
        out = tempfile.mktemp(suffix=".xlsx")
        try:
            ExcelReportGenerator().generate(
                [{"item": "X", "amount": 100.0}], out)
            wb = openpyxl.load_workbook(out)
            assert "Summary" in wb.sheetnames, \
                f"'Summary' not in {wb.sheetnames}"
        finally:
            if os.path.exists(out):
                os.unlink(out)

    def test_header_row_is_bold(self):
        """Header row cell at (row=1, col=1) has font.bold == True."""
        self._skip_unless_importable()
        import tempfile
        try:
            import openpyxl
            from xlsx.generator import ExcelReportGenerator
        except Exception as exc:
            pytest.skip(f"Cannot import: {exc}")
        out = tempfile.mktemp(suffix=".xlsx")
        try:
            ExcelReportGenerator().generate(
                [{"item": "X", "amount": 100.0}], out)
            wb = openpyxl.load_workbook(out)
            ws = wb.active
            assert ws.cell(row=1, column=1).font.bold is True, \
                "Header cell font.bold is not True"
        finally:
            if os.path.exists(out):
                os.unlink(out)

    def test_empty_data_creates_header_only_workbook(self):
        """generate([], tmp_path) creates workbook with only a header row."""
        self._skip_unless_importable()
        import tempfile
        try:
            import openpyxl
            from xlsx.generator import ExcelReportGenerator
        except Exception as exc:
            pytest.skip(f"Cannot import: {exc}")
        out = tempfile.mktemp(suffix=".xlsx")
        try:
            ExcelReportGenerator().generate([], out)
            wb = openpyxl.load_workbook(out)
            ws = wb.active
            assert ws.max_row == 1, \
                f"Expected 1 row (header only), got {ws.max_row}"
        finally:
            if os.path.exists(out):
                os.unlink(out)
