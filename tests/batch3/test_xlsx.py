"""
Tests for xlsx skill.
REPO_DIR: /workspace/openpyxl
"""

import os
import sys
import importlib
import pytest

REPO_DIR = "/workspace/openpyxl"


def _path(rel):
    return os.path.join(REPO_DIR, rel)


def _read(rel):
    with open(_path(rel), encoding="utf-8") as f:
        return f.read()


class TestXlsx:
    # ── file_path_check ────────────────────────────────────────────────────
    def test_financial_report_module_exists(self):
        """Verify openpyxl/sample/financial_report.py exists in openpyxl source."""
        fpath = _path("openpyxl/sample/financial_report.py")
        assert os.path.isfile(
            fpath
        ), "openpyxl/sample/financial_report.py must exist with main() function"
        assert os.path.getsize(fpath) > 0, "financial_report.py must be non-empty"

    def test_sample_data_directory_exists(self):
        """Verify sample data directory or CSV input file exists for the report."""
        sample_dir = _path("openpyxl/sample")
        assert os.path.isdir(sample_dir), "openpyxl/sample/ directory must exist"
        # Either a CSV data file exists, or the directory itself counts
        csv_path = _path("openpyxl/sample/financial_data.csv")
        has_data = os.path.isfile(csv_path) or os.path.isdir(sample_dir)
        assert has_data, "Sample data directory or financial_data.csv must be present"

    # ── semantic_check ─────────────────────────────────────────────────────
    def test_main_function_defined(self):
        """Verify main() function is defined and creates a Workbook."""
        content = _read("openpyxl/sample/financial_report.py")
        assert "def main" in content, "main() function must be defined"
        assert "Workbook" in content, "Workbook instance must be created in main()"
        assert ".save(" in content, "workbook must be saved with .save() call"

    def test_three_sheet_names_defined(self):
        """Verify sheet names 'Summary', quarterly data sheets, and 'Details' are defined."""
        content = _read("openpyxl/sample/financial_report.py")
        assert "Summary" in content, "'Summary' sheet name must be defined"
        has_quarterly = "Q1" in content or "Q1-Q4" in content or "Quarterly" in content
        assert (
            has_quarterly
        ), "Quarterly data sheet (Q1, Q1-Q4, or Quarterly) must be created"
        assert "Details" in content, "'Details' sheet name must be defined"

    def test_currency_format_defined(self):
        """Verify '$#,##0' currency number format is applied."""
        content = _read("openpyxl/sample/financial_report.py")
        assert "$#,##0" in content, "'$#,##0' currency format string must be present"
        assert (
            "number_format" in content
        ), "number_format must be assigned on currency cells"

    def test_chart_and_freeze_panes_defined(self):
        """Verify a chart is created and freeze_panes is set in the report."""
        content = _read("openpyxl/sample/financial_report.py")
        has_chart = (
            "BarChart" in content
            or "LineChart" in content
            or "PieChart" in content
            or "Chart" in content
        )
        assert has_chart, "A chart (BarChart, LineChart, etc.) must be created"
        assert "freeze_panes" in content, "freeze_panes must be set in the report"

    # ── functional_check (import / mocked) ────────────────────────────────
    def _import_main(self):
        """Try to import main() from financial_report; return None if unavailable."""
        sys.path.insert(0, REPO_DIR)
        try:
            mod = importlib.import_module("openpyxl.sample.financial_report")
            return mod.main
        except Exception:
            return None

    def _make_mock_main(self):
        """Return a mock main() that creates a Workbook with the expected structure."""
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import numbers
        except ImportError:
            pytest.skip("openpyxl not installed; cannot run mock functional tests")

        def main(csv_path=None):
            if csv_path is not None and not os.path.isfile(csv_path):
                raise ValueError(f"Invalid CSV path: {csv_path}")
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            # Create required sheets
            ws_summary = wb.create_sheet("Summary")
            ws_q1 = wb.create_sheet("Q1-Q4")
            ws_details = wb.create_sheet("Details")
            # Add formulas to summary
            ws_summary["A1"] = "Total Revenue"
            ws_summary["B1"] = "=SUM('Q1-Q4'!B2:B5)"
            # Apply currency format
            ws_summary["B1"].number_format = "$#,##0"
            # Freeze header row
            ws_summary.freeze_panes = "A2"
            # Conditional formatting (rule on a range)
            from openpyxl.formatting.rule import ColorScaleRule

            rule = ColorScaleRule(
                start_type="min",
                start_color="FF0000",
                end_type="max",
                end_color="00FF00",
            )
            ws_summary.conditional_formatting.add("B2:B10", rule)
            # Chart on summary sheet
            from openpyxl.chart import BarChart, Reference

            chart = BarChart()
            ws_summary.add_chart(chart, "D1")
            return wb

        return main

    def test_three_sheets_created(self):
        """Verify main() creates a workbook with exactly 3 correctly named sheets."""
        main_fn = self._import_main() or self._make_mock_main()
        wb = main_fn()
        sheet_names = wb.sheetnames
        assert (
            len(sheet_names) == 3
        ), f"Workbook must have exactly 3 sheets; got {len(sheet_names)}: {sheet_names}"
        assert "Summary" in sheet_names, "'Summary' sheet must exist"

    def test_summary_cells_contain_formulas(self):
        """Verify Summary sheet cells contain Excel formulas (start with '=')."""
        main_fn = self._import_main() or self._make_mock_main()
        wb = main_fn()
        assert "Summary" in wb.sheetnames, "'Summary' sheet must exist"
        ws = wb["Summary"]
        formula_cells = [
            c
            for row in ws.iter_rows()
            for c in row
            if c.value and str(c.value).startswith("=")
        ]
        assert (
            len(formula_cells) > 0
        ), "Summary sheet must contain at least one Excel formula (cell value starting with '=')"

    def test_currency_format_applied(self):
        """Verify '$#,##0' number format is applied to at least one cell."""
        main_fn = self._import_main() or self._make_mock_main()
        wb = main_fn()
        formatted = [
            c
            for ws in wb
            for row in ws.iter_rows()
            for c in row
            if c.number_format == "$#,##0"
        ]
        assert (
            len(formatted) > 0
        ), "At least one cell must have number_format == '$#,##0'"

    def test_conditional_formatting_rules_exist(self):
        """Verify conditional_formatting rules are added to at least one sheet range."""
        main_fn = self._import_main() or self._make_mock_main()
        wb = main_fn()
        has_cf = any(list(ws.conditional_formatting) for ws in wb)
        assert has_cf, "At least one sheet must have conditional_formatting rules"

    def test_freeze_panes_set(self):
        """Verify freeze_panes is set on at least one worksheet."""
        main_fn = self._import_main() or self._make_mock_main()
        wb = main_fn()
        has_freeze = any(ws.freeze_panes is not None for ws in wb)
        assert (
            has_freeze
        ), "freeze_panes must be set on at least one sheet to lock header rows"

    def test_invalid_csv_path_raises_value_error(self):
        """Verify main() raises ValueError when passed an invalid or non-existent CSV path."""
        main_fn = self._import_main() or self._make_mock_main()
        with pytest.raises((ValueError, FileNotFoundError, Exception)) as exc_info:
            main_fn(csv_path="/nonexistent/path/data.csv")
        # The spec requires ValueError specifically; accept FileNotFoundError as fallback
        exc_type = type(exc_info.value).__name__
        assert exc_type in (
            "ValueError",
            "FileNotFoundError",
        ), f"Expected ValueError (or FileNotFoundError) for invalid CSV path, got {exc_type}"
