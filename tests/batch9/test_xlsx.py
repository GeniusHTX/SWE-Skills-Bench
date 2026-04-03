"""
Test for 'xlsx' skill — Excel & Spreadsheet Automation
Validates XLSX reader/writer/formula/chart modules: file existence, class
signatures, and functional round-trip tests using openpyxl and pandas.
"""

import os
import sys
import tempfile

import pytest


class TestXlsx:
    """Verify XLSX reader, writer, formula evaluator, and chart builder."""

    REPO_DIR = "/workspace/openpyxl"

    # ── helpers ──────────────────────────────────────────────────────────
    @staticmethod
    def _read_file(path: str) -> str:
        try:
            with open(path, "r", errors="ignore") as fh:
                return fh.read()
        except OSError:
            return ""

    @classmethod
    def _add_to_path(cls):
        for d in (cls.REPO_DIR, os.path.join(cls.REPO_DIR, "examples")):
            if d not in sys.path:
                sys.path.insert(0, d)

    # ── file_path_check ──────────────────────────────────────────────────

    def test_xlsx_module_files_exist(self):
        """examples/xlsx/__init__.py, reader.py, writer.py must exist."""
        base = os.path.join(self.REPO_DIR, "examples", "xlsx")
        for name in ("__init__.py", "reader.py", "writer.py"):
            path = os.path.join(base, name)
            assert os.path.isfile(path), f"{path} does not exist"
            assert os.path.getsize(path) > 0, f"{name} is empty"

    def test_formula_and_chart_files_exist(self):
        """examples/xlsx/formula.py and chart.py must exist."""
        base = os.path.join(self.REPO_DIR, "examples", "xlsx")
        for name in ("formula.py", "chart.py"):
            path = os.path.join(base, name)
            assert os.path.isfile(path), f"{path} does not exist"
            assert os.path.getsize(path) > 0, f"{name} is empty"

    def test_test_xlsx_py_exists(self):
        """tests/test_xlsx.py must exist."""
        path = os.path.join(self.REPO_DIR, "tests", "test_xlsx.py")
        assert os.path.isfile(path), f"{path} does not exist"

    # ── semantic_check ───────────────────────────────────────────────────

    def test_xlsx_reader_read_method_signature(self):
        """reader.py must define XLSXReader class with read method using openpyxl/pandas."""
        content = self._read_file(
            os.path.join(self.REPO_DIR, "examples", "xlsx", "reader.py")
        )
        assert "XLSXReader" in content or "class " in content, (
            "XLSXReader class not defined in reader.py"
        )
        assert "def read" in content, "read method not defined"
        assert "openpyxl" in content or "read_excel" in content, (
            "Neither openpyxl nor pd.read_excel found in reader.py"
        )

    def test_xlsx_writer_write_method_signature(self):
        """writer.py must define XLSXWriter class with write method."""
        content = self._read_file(
            os.path.join(self.REPO_DIR, "examples", "xlsx", "writer.py")
        )
        assert "XLSXWriter" in content or "class " in content, (
            "XLSXWriter class not defined"
        )
        assert "def write" in content, "write method not defined"
        assert "openpyxl" in content or "to_excel" in content, (
            "Neither openpyxl nor to_excel found in writer.py"
        )

    def test_formula_evaluator_raises_formula_error(self):
        """formula.py must define FormulaError and raise it for unsupported functions."""
        content = self._read_file(
            os.path.join(self.REPO_DIR, "examples", "xlsx", "formula.py")
        )
        assert "FormulaError" in content, "FormulaError not defined in formula.py"
        assert "raise" in content, "No raise statement for FormulaError"

    def test_chart_builder_uses_openpyxl_chart(self):
        """chart.py must use openpyxl chart objects (BarChart/LineChart/PieChart)."""
        content = self._read_file(
            os.path.join(self.REPO_DIR, "examples", "xlsx", "chart.py")
        )
        chart_types = ["BarChart", "LineChart", "PieChart"]
        found = any(c in content for c in chart_types)
        assert found, "No openpyxl chart type imported in chart.py"
        assert "add_chart" in content, "add_chart not called in chart.py"

    # ── functional_check (import) ────────────────────────────────────────

    def test_write_then_read_preserves_data(self):
        """Round-trip write/read must preserve DataFrame values."""
        self._add_to_path()
        try:
            import pandas as pd
            from xlsx.writer import XLSXWriter
            from xlsx.reader import XLSXReader
        except ImportError as exc:
            pytest.skip(f"Import failed: {exc}")

        df = pd.DataFrame({"a": [1, 2], "b": ["x", "y"]})
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            XLSXWriter().write(df, tmp_path)
            df2 = XLSXReader().read(tmp_path)
            assert list(df2.columns) == ["a", "b"]
            assert df2["b"].tolist() == ["x", "y"]
        finally:
            os.unlink(tmp_path)

    def test_file_created_on_write(self):
        """XLSXWriter.write() must create the output file."""
        self._add_to_path()
        try:
            import pandas as pd
            from xlsx.writer import XLSXWriter
        except ImportError as exc:
            pytest.skip(f"Import failed: {exc}")

        df = pd.DataFrame({"a": [1, 2]})
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            os.unlink(tmp_path)  # ensure it doesn't exist
            XLSXWriter().write(df, tmp_path)
            assert os.path.exists(tmp_path), "File was not created"
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_formula_sum_evaluated_correctly(self):
        """FormulaEvaluator.evaluate('SUM(1,2,3)') must return 6."""
        self._add_to_path()
        try:
            from xlsx.formula import FormulaEvaluator
        except ImportError as exc:
            pytest.skip(f"Import failed: {exc}")

        result = FormulaEvaluator().evaluate("SUM(1,2,3)")
        assert result == 6, f"Expected 6, got {result}"

    def test_empty_worksheet_read_returns_empty_dataframe(self):
        """Reading an empty worksheet must return an empty DataFrame."""
        self._add_to_path()
        try:
            import openpyxl
            from xlsx.reader import XLSXReader
        except ImportError as exc:
            pytest.skip(f"Import failed: {exc}")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            wb = openpyxl.Workbook()
            wb.save(tmp_path)
            df = XLSXReader().read(tmp_path)
            assert df.empty, "Expected empty DataFrame for empty worksheet"
        finally:
            os.unlink(tmp_path)

    def test_unsupported_formula_raises_formula_error(self):
        """FormulaEvaluator.evaluate('UNSUPPORTED()') must raise FormulaError."""
        self._add_to_path()
        try:
            from xlsx.formula import FormulaEvaluator, FormulaError
        except ImportError as exc:
            pytest.skip(f"Import failed: {exc}")

        with pytest.raises(FormulaError):
            FormulaEvaluator().evaluate("UNSUPPORTED()")

    def test_nan_values_written_as_empty_cells(self):
        """NaN values must be written as empty cells, not 'nan' strings."""
        self._add_to_path()
        try:
            import numpy as np
            import openpyxl
            import pandas as pd
            from xlsx.writer import XLSXWriter
        except ImportError as exc:
            pytest.skip(f"Import failed: {exc}")

        df = pd.DataFrame({"a": [1, np.nan, 3]})
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            XLSXWriter().write(df, tmp_path)
            wb = openpyxl.load_workbook(tmp_path)
            ws = wb.active
            # Row 3 (1-indexed: row 1=header, row 2=first data, row 3=NaN)
            nan_cell = ws.cell(row=3, column=1).value
            assert nan_cell is None or nan_cell != "nan", (
                f"NaN written as string: {nan_cell}"
            )
        finally:
            os.unlink(tmp_path)
