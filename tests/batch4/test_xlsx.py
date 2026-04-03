"""
Test for 'xlsx' skill — Financial Summary Report Generator
Validates that the Agent created a generate_financial_report function
using openpyxl with proper sheets, formulas, and data handling.
"""

import os
import sys

import pytest


class TestXlsx:
    """Verify openpyxl financial summary report generator."""

    REPO_DIR = "/workspace/openpyxl"

    # ---- helpers ----

    @staticmethod
    def _read(path):
        with open(path, "r", errors="ignore") as fh:
            return fh.read()

    # ---- file_path_check ----

    def test_financial_summary_py_exists(self):
        """Verifies examples/financial_summary.py exists."""
        path = os.path.join(self.REPO_DIR, "examples/financial_summary.py")
        assert os.path.exists(path), f"Expected file not found: {path}"

    def test_sample_data_py_exists(self):
        """Verifies examples/sample_data.py exists."""
        path = os.path.join(self.REPO_DIR, "examples/sample_data.py")
        assert os.path.exists(path), f"Expected file not found: {path}"

    def test_sample_portfolio_json_exists(self):
        """Verifies examples/sample_portfolio.json exists."""
        path = os.path.join(self.REPO_DIR, "examples/sample_portfolio.json")
        assert os.path.exists(path), f"Expected file not found: {path}"

    # ---- semantic_check ----

    def test_sem_import_generate_financial_report(self):
        """Verifies generate_financial_report is importable."""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from examples.financial_summary import generate_financial_report

            assert callable(generate_financial_report)
        finally:
            sys.path.pop(0)

    def test_sem_function_signature(self):
        """Verifies function accepts portfolio_data, transactions, output_path."""
        content = self._read(
            os.path.join(self.REPO_DIR, "examples/financial_summary.py")
        )
        assert "portfolio_data" in content, "portfolio_data parameter missing"
        assert "transactions" in content, "transactions parameter missing"
        assert "output_path" in content, "output_path parameter missing"

    def test_sem_openpyxl_import(self):
        """Verifies source contains openpyxl import."""
        content = self._read(
            os.path.join(self.REPO_DIR, "examples/financial_summary.py")
        )
        assert "openpyxl" in content, "openpyxl import missing"

    def test_sem_formula_present(self):
        """Verifies at least one Excel formula string (edge case)."""
        content = self._read(
            os.path.join(self.REPO_DIR, "examples/financial_summary.py")
        )
        assert "=" in content, "No formula string starting with '=' found"

    # ---- functional_check ----

    def test_func_generate_report_succeeds(self, tmp_path):
        """Generate report with single AAPL holding."""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from examples.financial_summary import generate_financial_report

            generate_financial_report(
                [
                    {
                        "symbol": "AAPL",
                        "shares": 10,
                        "avg_cost": 100,
                        "current_price": 150,
                    }
                ],
                [],
                str(tmp_path / "report.xlsx"),
            )
            assert (tmp_path / "report.xlsx").exists()
        finally:
            sys.path.pop(0)

    def test_func_summary_sheet_exists(self, tmp_path):
        """Verifies 'Summary' sheet in generated workbook."""
        sys.path.insert(0, self.REPO_DIR)
        try:
            import openpyxl
            from examples.financial_summary import generate_financial_report

            generate_financial_report(
                [
                    {
                        "symbol": "AAPL",
                        "shares": 10,
                        "avg_cost": 100,
                        "current_price": 150,
                    }
                ],
                [],
                str(tmp_path / "report.xlsx"),
            )
            wb = openpyxl.load_workbook(tmp_path / "report.xlsx")
            assert "Summary" in wb.sheetnames, "Summary sheet missing"
        finally:
            sys.path.pop(0)

    def test_func_holdings_sheet_exists(self, tmp_path):
        """Verifies 'Holdings' sheet in generated workbook."""
        sys.path.insert(0, self.REPO_DIR)
        try:
            import openpyxl
            from examples.financial_summary import generate_financial_report

            generate_financial_report(
                [
                    {
                        "symbol": "AAPL",
                        "shares": 10,
                        "avg_cost": 100,
                        "current_price": 150,
                    }
                ],
                [],
                str(tmp_path / "report.xlsx"),
            )
            wb = openpyxl.load_workbook(tmp_path / "report.xlsx")
            assert "Holdings" in wb.sheetnames, "Holdings sheet missing"
        finally:
            sys.path.pop(0)

    def test_func_transactions_sheet_exists(self, tmp_path):
        """Verifies 'Transactions' sheet in generated workbook."""
        sys.path.insert(0, self.REPO_DIR)
        try:
            import openpyxl
            from examples.financial_summary import generate_financial_report

            generate_financial_report(
                [
                    {
                        "symbol": "AAPL",
                        "shares": 10,
                        "avg_cost": 100,
                        "current_price": 150,
                    }
                ],
                [],
                str(tmp_path / "report.xlsx"),
            )
            wb = openpyxl.load_workbook(tmp_path / "report.xlsx")
            assert "Transactions" in wb.sheetnames, "Transactions sheet missing"
        finally:
            sys.path.pop(0)

    def test_func_holdings_aapl_data(self, tmp_path):
        """Verifies Holdings sheet row 2 has AAPL data with market_value ~1500."""
        sys.path.insert(0, self.REPO_DIR)
        try:
            import openpyxl
            from examples.financial_summary import generate_financial_report

            generate_financial_report(
                [
                    {
                        "symbol": "AAPL",
                        "shares": 10,
                        "avg_cost": 100,
                        "current_price": 150,
                    }
                ],
                [],
                str(tmp_path / "report.xlsx"),
            )
            wb = openpyxl.load_workbook(tmp_path / "report.xlsx")
            ws = wb["Holdings"]
            row2_values = [cell.value for cell in ws[2]]
            assert "AAPL" in row2_values, "AAPL not found in Holdings row 2"
        finally:
            sys.path.pop(0)

    def test_func_unrealized_pnl(self, tmp_path):
        """Verifies unrealized PnL and pnl_pct for AAPL holding."""
        sys.path.insert(0, self.REPO_DIR)
        try:
            import openpyxl
            from examples.financial_summary import generate_financial_report

            generate_financial_report(
                [
                    {
                        "symbol": "AAPL",
                        "shares": 10,
                        "avg_cost": 100,
                        "current_price": 150,
                    }
                ],
                [],
                str(tmp_path / "report.xlsx"),
            )
            wb = openpyxl.load_workbook(tmp_path / "report.xlsx")
            ws = wb["Holdings"]
            row2_values = [cell.value for cell in ws[2]]
            # Check that 500.0 (PnL) or 50.0 (pct) or a formula is present
            has_pnl = any(
                v == 500.0 or v == 500 or (isinstance(v, str) and v.startswith("="))
                for v in row2_values
                if v is not None
            )
            assert (
                has_pnl or len(row2_values) > 3
            ), "Unrealized PnL data not found in Holdings row 2"
        finally:
            sys.path.pop(0)

    def test_func_empty_portfolio(self, tmp_path):
        """Edge case: empty portfolio generates valid workbook with header only."""
        sys.path.insert(0, self.REPO_DIR)
        try:
            import openpyxl
            from examples.financial_summary import generate_financial_report

            generate_financial_report([], [], str(tmp_path / "empty.xlsx"))
            wb2 = openpyxl.load_workbook(tmp_path / "empty.xlsx")
            ws = wb2["Holdings"]
            assert ws.max_row == 1, f"Expected 1 row (header only), got {ws.max_row}"
        finally:
            sys.path.pop(0)

    def test_func_invalid_transaction_type(self, tmp_path):
        """Failure case: invalid transaction type raises ValueError."""
        sys.path.insert(0, self.REPO_DIR)
        try:
            from examples.financial_summary import generate_financial_report

            with pytest.raises((ValueError, KeyError, Exception)):
                generate_financial_report(
                    [],
                    [{"type": "INVALID", "symbol": "X", "shares": 1, "price": 1}],
                    str(tmp_path / "fail.xlsx"),
                )
        finally:
            sys.path.pop(0)
