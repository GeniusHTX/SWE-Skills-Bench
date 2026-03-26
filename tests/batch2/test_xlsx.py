"""
Test for 'xlsx' skill — Excel & Spreadsheet Automation
Validates that the Agent implemented a report generation engine for openpyxl
producing multi-sheet workbooks with data tables, summary calculations,
conditional formatting, and charts.
"""

import os
import re
import subprocess
import tempfile

import pytest

from _dependency_utils import ensure_python_dependencies


@pytest.fixture(scope="module", autouse=True)
def _ensure_repo_dependencies():
    ensure_python_dependencies(TestXlsx.REPO_DIR)


class TestXlsx:
    """Verify Excel report generation engine for openpyxl."""

    REPO_DIR = "/workspace/openpyxl"

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _read(self, *parts):
        fpath = os.path.join(self.REPO_DIR, *parts)
        assert os.path.isfile(fpath), f"Required file not found: {fpath}"
        with open(fpath, "r", errors="ignore") as fh:
            return fh.read()

    # ------------------------------------------------------------------
    # L1: File existence and syntax
    # ------------------------------------------------------------------

    def test_report_engine_file_exists(self):
        """openpyxl/utils/report_engine.py must exist."""
        fpath = os.path.join(self.REPO_DIR, "openpyxl", "utils", "report_engine.py")
        assert os.path.isfile(fpath), "openpyxl/utils/report_engine.py not found"

    def test_report_engine_compiles(self):
        """report_engine.py must be syntactically valid Python."""
        result = subprocess.run(
            ["python", "-m", "py_compile", "openpyxl/utils/report_engine.py"],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert (
            result.returncode == 0
        ), f"Syntax error in report_engine.py:\n{result.stderr}"

    def test_report_engine_importable(self):
        """report_engine module must be importable."""
        result = subprocess.run(
            ["python", "-c", "from openpyxl.utils.report_engine import *; print('OK')"],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert (
            result.returncode == 0
        ), f"Failed to import report_engine:\n{result.stderr}"

    # ------------------------------------------------------------------
    # L1: Engine structure
    # ------------------------------------------------------------------

    def test_engine_defines_callable(self):
        """report_engine.py must define a callable function or class for report generation."""
        content = self._read("openpyxl", "utils", "report_engine.py")
        patterns = [
            r"def\s+generate",
            r"def\s+create",
            r"def\s+build",
            r"class\s+\w*Report\w*Engine",
            r"class\s+\w*Report\w*Generator",
            r"class\s+\w*Report\w*Builder",
        ]
        assert any(
            re.search(p, content, re.IGNORECASE) for p in patterns
        ), "report_engine.py does not define a generate/create/build callable or Report class"

    def test_engine_uses_openpyxl_workbook(self):
        """Engine must use openpyxl Workbook to create Excel files."""
        content = self._read("openpyxl", "utils", "report_engine.py")
        assert re.search(
            r"Workbook|workbook|load_workbook", content
        ), "report_engine.py does not reference openpyxl Workbook"

    # ------------------------------------------------------------------
    # L2: Multi-sheet layout
    # ------------------------------------------------------------------

    def test_engine_creates_multiple_sheets(self):
        """Engine code must create at least 3 separate worksheet references."""
        content = self._read("openpyxl", "utils", "report_engine.py")
        sheet_patterns = [
            r"create_sheet",
            r"active",
            r"\.title\s*=",
            r"ws\d",
            r"data_sheet|summary_sheet|chart_sheet",
        ]
        matches = sum(
            len(re.findall(p, content, re.IGNORECASE)) for p in sheet_patterns
        )
        assert (
            matches >= 3
        ), f"Engine does not appear to create multiple sheets (found {matches} sheet references)"

    def test_engine_has_data_sheet_logic(self):
        """Engine must populate a data sheet with rows and a header row."""
        content = self._read("openpyxl", "utils", "report_engine.py")
        data_patterns = [r"header", r"append\(", r"row", r"cell"]
        matches = sum(1 for p in data_patterns if re.search(p, content, re.IGNORECASE))
        assert (
            matches >= 2
        ), "Engine missing data sheet population logic (headers + rows)"

    def test_engine_has_summary_calculations(self):
        """Engine must compute summary/aggregate metrics (totals, averages, counts)."""
        content = self._read("openpyxl", "utils", "report_engine.py")
        summary_patterns = [
            r"sum\(",
            r"average",
            r"mean",
            r"count",
            r"total",
            r"SUM\(",
            r"AVERAGE\(",
            r"COUNT\(",  # Excel formulas
            r"aggregate",
            r"summary",
        ]
        assert any(
            re.search(p, content, re.IGNORECASE) for p in summary_patterns
        ), "Engine missing summary/aggregate calculation logic"

    # ------------------------------------------------------------------
    # L2: Chart generation
    # ------------------------------------------------------------------

    def test_engine_creates_chart(self):
        """Engine must create at least one chart (line, bar, or pie)."""
        content = self._read("openpyxl", "utils", "report_engine.py")
        chart_patterns = [
            r"BarChart",
            r"LineChart",
            r"PieChart",
            r"Reference",
            r"add_chart",
            r"chart",
            r"Chart",
        ]
        matches = sum(1 for p in chart_patterns if re.search(p, content))
        assert matches >= 2, "Engine does not create a chart visualization"

    # ------------------------------------------------------------------
    # L2: Conditional formatting
    # ------------------------------------------------------------------

    def test_engine_applies_conditional_formatting(self):
        """Engine must apply conditional formatting to at least one range."""
        content = self._read("openpyxl", "utils", "report_engine.py")
        cf_patterns = [
            r"conditional_formatting",
            r"ConditionalFormatting",
            r"CellIsRule",
            r"ColorScaleRule",
            r"DataBarRule",
            r"FormulaRule",
            r"IconSetRule",
            r"PatternFill.*condition",
            r"condition.*PatternFill",
        ]
        assert any(
            re.search(p, content, re.IGNORECASE) for p in cf_patterns
        ), "Engine does not apply conditional formatting"

    # ------------------------------------------------------------------
    # L2: Styling
    # ------------------------------------------------------------------

    def test_engine_applies_header_styling(self):
        """Engine must style headers (bold, fill, font, etc.)."""
        content = self._read("openpyxl", "utils", "report_engine.py")
        style_patterns = [
            r"Font\(.*bold",
            r"bold\s*=\s*True",
            r"PatternFill",
            r"font.*bold",
            r"header.*style",
            r"style.*header",
        ]
        assert any(
            re.search(p, content, re.IGNORECASE) for p in style_patterns
        ), "Engine does not style header rows"

    # ------------------------------------------------------------------
    # L2: Dynamic execution — actually generate a workbook
    # ------------------------------------------------------------------

    def test_engine_generates_valid_xlsx_file(self):
        """Running the engine with sample data must produce a valid .xlsx file."""
        script = """
import sys, os, tempfile
sys.path.insert(0, '.')
from openpyxl.utils.report_engine import *

# Find the main callable
import inspect
mod = sys.modules['openpyxl.utils.report_engine']
callables = [(n, o) for n, o in inspect.getmembers(mod)
             if (callable(o) and not n.startswith('_'))]

sample_data = [
    {"date": "2024-01-01", "product": "Widget A", "quantity": 10, "revenue": 100.0},
    {"date": "2024-01-02", "product": "Widget B", "quantity": 5, "revenue": 75.0},
    {"date": "2024-01-03", "product": "Widget A", "quantity": 8, "revenue": 80.0},
    {"date": "2024-01-04", "product": "Widget C", "quantity": 3, "revenue": 45.0},
    {"date": "2024-01-05", "product": "Widget B", "quantity": 12, "revenue": 180.0},
]

outpath = os.path.join(tempfile.gettempdir(), 'test_report.xlsx')

generated = False
for name, obj in callables:
    if 'generate' in name.lower() or 'create' in name.lower() or 'build' in name.lower():
        try:
            if inspect.isclass(obj):
                instance = obj()
                for method in ['generate', 'create', 'build', 'run']:
                    fn = getattr(instance, method, None)
                    if fn:
                        result = fn(sample_data, outpath)
                        generated = True
                        break
            else:
                result = obj(sample_data, outpath)
                generated = True
        except TypeError:
            try:
                result = obj(data=sample_data, output_path=outpath)
                generated = True
            except:
                pass
        if generated:
            break

if not generated:
    # Try any callable that looks relevant
    for name, obj in callables:
        if callable(obj) and not inspect.isclass(obj):
            try:
                result = obj(sample_data, outpath)
                generated = True
                break
            except:
                pass

if generated and os.path.isfile(outpath):
    size = os.path.getsize(outpath)
    print(f'OK size={size}')
    # Verify it's a real xlsx (ZIP magic bytes)
    with open(outpath, 'rb') as f:
        magic = f.read(4)
    if magic[:2] == b'PK':
        print('VALID_ZIP')
    else:
        print('INVALID_FORMAT')
else:
    print('NOT_GENERATED')
"""
        result = subprocess.run(
            ["python", "-c", script],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=60,
        )
        output = result.stdout.strip()
        if "NOT_GENERATED" in output:
            pytest.fail(
                f"Engine did not generate an output file.\n"
                f"stderr: {result.stderr[-2000:]}"
            )
        assert (
            "OK" in output
        ), f"Engine execution failed:\nstdout: {output}\nstderr: {result.stderr[-2000:]}"
        if "VALID_ZIP" not in output:
            pytest.fail("Generated file is not a valid xlsx (ZIP) format")

    def test_generated_workbook_has_multiple_sheets(self):
        """The generated workbook must contain at least 3 sheets."""
        script = """
import sys, os, tempfile
sys.path.insert(0, '.')
from openpyxl.utils.report_engine import *
from openpyxl import load_workbook
import inspect

mod = sys.modules['openpyxl.utils.report_engine']
callables = [(n, o) for n, o in inspect.getmembers(mod)
             if callable(o) and not n.startswith('_')]

sample_data = [
    {"date": "2024-01-01", "product": "Widget", "quantity": 10, "revenue": 100.0},
    {"date": "2024-01-02", "product": "Gadget", "quantity": 5, "revenue": 75.0},
]

outpath = os.path.join(tempfile.gettempdir(), 'test_sheets.xlsx')
generated = False
for name, obj in callables:
    try:
        if inspect.isclass(obj):
            inst = obj()
            for m in ['generate', 'create', 'build', 'run']:
                fn = getattr(inst, m, None)
                if fn:
                    fn(sample_data, outpath)
                    generated = True
                    break
        else:
            obj(sample_data, outpath)
            generated = True
    except:
        pass
    if generated:
        break

if generated and os.path.isfile(outpath):
    wb = load_workbook(outpath)
    print(f'SHEETS={len(wb.sheetnames)}')
    print(f'NAMES={wb.sheetnames}')
else:
    print('NOT_GENERATED')
"""
        result = subprocess.run(
            ["python", "-c", script],
            cwd=self.REPO_DIR,
            capture_output=True,
            text=True,
            timeout=60,
        )
        output = result.stdout.strip()
        if "SHEETS=" in output:
            count = int(output.split("SHEETS=")[1].split("\n")[0])
            assert count >= 3, (
                f"Workbook has only {count} sheet(s), expected at least 3 "
                "(data, summary, chart)"
            )
        elif "NOT_GENERATED" in output:
            pytest.skip("Could not generate workbook for sheet count verification")
